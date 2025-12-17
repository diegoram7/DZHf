import marimo

__generated_with = "0.18.4"
app = marimo.App(width="medium", layout_file="layouts/main.grid.json")


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    #Eocene to Oligocene source-to-sink dynamics in Panama using U-Pb-Hf in zircons
    """)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ##Import libraries
    """)
    return


@app.cell
def _():
    import pandas as pd
    import marimo as mo
    import numpy as np
    import plotly.express as px
    import plotly.graph_objects as go
    import openpyxl
    import quak
    return go, mo, np, pd, px, quak


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ##Create and adjust dataframe
    """)
    return


@app.cell
def _():
    DATA_LOCATION = "https://diegoram7.github.io/DZHf/public/data/Hf.csv"
    return (DATA_LOCATION,)


@app.cell
def _(DATA_LOCATION, pd):
    #Create dataframe 
    df = pd.read_csv(DATA_LOCATION, sep= ";", decimal=",", thousands=".").rename(columns={"1s_error.1":"1s_error-1"})

    #Ajuste de nombre de columna Sample para cálculos y gráficos
    df[["sampleid", "number"]] = df['Sample'].str.split('_', expand=True, n=1) 

    #Conversión de unidades de Ga a Ma
    df["t(Ma)"] = df["t(Ga)"]*1000

    return (df,)


@app.cell
def _(df, mo, quak):
    widget = mo.ui.anywidget(quak.Widget(df))
    widget
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md("""
    ## Calculate $\epsilon$-Hafnium and error propagation
    """)
    return


@app.cell
def _():
    # Lu and Hf constants
    # Vervoort et al. 2018
    Hf_DM = 0.283225
    Lu_DM = 0.0383

    # Vervoort & Blichert-Toft 1999
    Hf_CHUR = 0.282772
    Lu_CHUR = 0.0332

    #Söderlund et al. (2024)
    lam_Lu = 1.867e-11 # yr^-1

    #Griffin et al. (2004)
    Lu_crust = 0.015
    return Hf_CHUR, Hf_DM, Lu_CHUR, Lu_DM, Lu_crust, lam_Lu


@app.cell
def _(Hf_CHUR, Lu_CHUR, df, lam_Lu, np):
    # calculate eHfi and propagate errors 

    # Columnas de interés
    t = df["t(Ga)"] * 1e9           #t (in years)
    Lu176_Hf177 = df["176Lu_177Hf"] 
    Hf176_Hf177 = df["176Hf_177Hf"]

    # Calculate eHfi
    decaimiento = np.exp( lam_Lu * t ) - 1
    _numerador = Hf176_Hf177 - Lu176_Hf177 * decaimiento
    CHUR_t = Hf_CHUR - Lu_CHUR * decaimiento 
    df["ehf"] = 10_000 * (_numerador / CHUR_t - 1)

    #Propagate errors at 2s
    two_sigma = 2*(pow(10, 4) * (df["1s_error-1"] / CHUR_t))
    df["2s"] = two_sigma
    return Hf176_Hf177, decaimiento


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ##Filter outliers and high errors
    """)
    return


@app.cell
def _(df):
    #Calculate q1 and q3 from age data 
    q1, q3 = df["t(Ga)"].quantile([.05, .95])
    IRQ = q3 - q1
    print(q1, q3)

    #Separate outliers and high uncertainty data
    df_good = df[
        mask := (
            df["t(Ga)"].between(q1 - IRQ, q3 + IRQ)
            & (df["2s"] < 2.0)
        )
    ]
    df_bad = df[-mask] #Outliers dataframe

    print(f"Size original: {df.shape}")
    print(f"Size filtered: {df_good.shape}")
    print(f"Ratio: {df_good.shape[0]/df.shape[0]}")
    return (df_good,)


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ##Calculate DM line for time segment
    """)
    return


@app.cell
def _(df_good):
    x_min = df_good['t(Ma)'].min()
    x_max = df_good['t(Ma)'].max()


    def chur_y(x):
        return 0

    chur_xs = (x_min, x_max)
    chur_ys = (chur_y(x_min), chur_y(x_max))

    def dm_y(x):
        y = -0.004 * (x) + 18
        return y

    dm_xs = (x_min, x_max)
    dm_ys = (dm_y(x_min), dm_y(x_max))
    return chur_xs, chur_ys, dm_xs, dm_ys


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ##$\epsilon$-Hafnio  vs Age (Ma)
    """)
    return


@app.cell
def _(chur_xs, chur_ys, df_good, dm_xs, dm_ys, go, px):
    #Create scatter figure
    fig = px.scatter(
        df_good,
        x="t(Ma)",
        y="ehf",
        color ="sampleid",
        error_y="2s",
        marginal_x="box",
        marginal_y="violin",
        title="<b>ε-Hf vs. Age (Ma)<b>")

    fig.update_layout(
        xaxis = dict(
            showline=True,
            showgrid=False,
            showticklabels=True,
            ticks="outside"
        ),

        yaxis = dict(
            showline=True,
            showgrid=False,
            showticklabels=True,
            ticks="outside"
        ),
        plot_bgcolor="white",
        paper_bgcolor="white",
        yaxis_title_text="<b>εHf</b>",
        xaxis_title_text="<b>time(Ma)</b>",
        legend=dict(
            title="<b>Legend<b>" 
        )
    )

    #Posible igneous sources
    #Parita pluton
    fig.add_trace(
        go.Scatter(
            x=[36, 37, 37, 36, 36],
            y=[-10, -10, 17, 17, -10],
            line=dict(color="black", width=0),
            fill="toself",
            fillcolor="rgba(255, 0, 0, 0.1)",
            name="Parita pluton"
        )
    )

    #Mamoní pluton
    fig.add_trace(
        go.Scatter(
            x=[39, 49, 49, 39, 39],
            y=[-10, -10, 17, 17, -10],
            line=dict(color="black", width=0),
            fill="toself",
            fillcolor="rgba(150, 0, 200, 0.2)",
            name="Mamoni pluton"
        )
    )

    #Valle Rico pluton
    fig.add_trace(
        go.Scatter(
            x=[48, 49, 49, 48, 48],
            y=[-10, -10, 17, 17, -10],
            line=dict(color="black", width=0),
            fill="toself",
            fillcolor="rgba(0, 200, 0, 0.25)",
            name="Valle Rico pluton"
        )
    )

    #Cerro Azul pluton
    fig.add_trace(
        go.Scatter(
            x=[54, 59, 59, 54, 54],
            y=[-10, -10, 17, 17, -10],
            line=dict(color="black", width=0),
            fill="toself",
            fillcolor="rgba(255, 140, 0, 0.3)",
            name="Cerro Azul pluton"
        )
    )

    #Cerro Montuoso pluton
    fig.add_trace(
        go.Scatter(
            x=[66, 67, 67, 66, 66],
            y=[-10, -10, 17, 17, -10],
            line=dict(color="black", width=0),
            fill="toself",
            fillcolor="rgba(0, 102, 255, 0.15)",
            name="Cerro Montuoso pluton"
        )
    )

    #DM and CHUR lines
    fig.add_trace(
         go.Scatter(
             x=dm_xs,
             y=dm_ys,
             name="DM (Depleted Mantle)",
             line=dict(color="Crimson"))
    )
    fig.add_scatter(name="CHUR (Chondritic uniform reservoir)", x=chur_xs, y=chur_ys)

    #DM and CHUR annotations
    fig.add_annotation(
        text="DM",
        x=dm_xs[0]+15,
        y=dm_ys[0]+1,
        showarrow=False
    )

    fig.add_annotation(
        text="CHUR",
        x=chur_xs[0]+15,
        y=chur_ys[0]-1,
        showarrow=False
    )

    #Zircon sources
    fig.add_trace(
        go.Scatter(
            x=[35, 70, 70, 35, 35],
            y=[5, 5, 15, 15, 5],
            line=dict(color="black", width=0),
            fill="toself",
            fillcolor="rgba(160, 160, 160, 0.2)",
            name="Juvenil",
            visible="legendonly"
        )
    )

    fig.add_trace(
        go.Scatter(
            x=[35, 70, 70, 35, 35],
            y=[-20, -20, -5, -5, -20],
            line=dict(color="black", width=0),
            fill="toself",
            fillcolor="rgba(160, 160, 160, 0.2)",
            name="Corteza reciclada",
            visible="legendonly"
        )
    )
    fig
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ## Hf data results
    """)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    Initial εHf value ranges (only for values with 2sigma < 2.0):

    - Magmatic zircons:
          Cerro Montuoso Pluton (060072): (+4.9 to +9.0) Parita Pluton (060065 - 060067): (+4.3 to +11.0)
    - Detrital zircons:
          Diablo River Sand (070242): (+3.7 to +13.1), with an outlier of -3.2
          Cobachón Formation (300351): (+2.6 to +11.6)
          Gatuncillo Formation (300339): (+7.1 to +15.4)
          Mamoni River Sand (300334): (+8.2 to 14.0)

    The results obtained indicate that the magmatic system formed during ca. 70-35 Ma is characterized by a juvenile signature, derived from mantle magmatism, without input from ancient continental crust.
    """)
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    ##Two‑stage crustal model age (Para circones antiguos, no valido para Panamá)
    """)
    return


@app.cell
def _(Hf176_Hf177, Hf_DM, Lu_DM, Lu_crust, decaimiento, lam_Lu, np):
    _numerador = Hf176_Hf177 + Lu_crust * decaimiento - Hf_DM
    _denominador = Lu_crust - Lu_DM

    #print(_numerador/_denominador)

    tdm2 = (1 / lam_Lu) * np.log(_numerador/_denominador + 1)
    tdm2
    return


@app.cell(hide_code=True)
def _(mo):
    mo.md(r"""
    #Detrital zircons analysis
    """)
    return


@app.cell
def _(DATA_LOCATION, pd):
    #Create dataframe2 
    df2 = pd.read_excel(DATA_LOCATION, sheet_name="ZrUPb")
    return


if __name__ == "__main__":
    app.run()
